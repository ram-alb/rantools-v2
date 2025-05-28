from io import BytesIO
from typing import Optional, Tuple

import openpyxl


def generate_bulk_template(technology: str, parameter: str) -> bytes:
    """Generate an Excel template for bulk configuration."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{technology}_{parameter}"
    ws.append(["cell", parameter])
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def validate_uploaded_template(
    template: BytesIO,
    expected_param: str,
) -> Tuple[bool, Optional[str]]:
    """Validate the uploaded Excel template for bulk configuration."""
    wb = openpyxl.load_workbook(template)

    # 1. check if the file contains exactly one sheet
    if len(wb.sheetnames) != 1:
        return False, "The file must contain exactly one sheet."
    ws = wb.active

    # 2. check if the first row contains the expected headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected_headers = ["cell", expected_param]
    if headers != expected_headers:
        return False, f"Columns must be: {expected_headers}"

    # 3. check if there is at least one data row (second row exists)
    second_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if not second_row:
        return False, "There must be at least one data row."

    return True, None
