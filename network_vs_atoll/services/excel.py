from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def fill_excel(node, diffs):
    """
    Fill excel file with inconsistencies.

    Args:
        node (str): a node name
        diffs (list): a list of dicts with parameters diff

    Returns:
        str: a report path
    """
    work_book = Workbook()
    sheet = work_book.active

    columns = list(diffs[0].keys())

    for col in columns:
        col_number = columns.index(col) + 1
        current_cell = sheet.cell(row=1, column=col_number)
        current_cell.value = col
        current_cell.alignment = Alignment(horizontal='center')
        sheet.column_dimensions[get_column_letter(col_number)].width = 30

    row = 2
    for diff in diffs:
        for column in columns:
            current_cell = sheet.cell(row=row, column=columns.index(column) + 1)
            current_cell.value = diff[column]
            current_cell.alignment = Alignment(horizontal='center')
        row += 1

    report_path = f'{node}.xlsx'
    work_book.save(report_path)
    return report_path
