from tempfile import NamedTemporaryFile
from typing import List

from openpyxl import Workbook  # type: ignore

from bts_files.services.atoll.main import CellRowFactory
from bts_files.services.filter_cells import AllTechPolygon


def _create_sheet(work_book: Workbook, technology: str, tech_data: List[CellRowFactory]) -> None:
    """Create a new sheet in the given workbook and populates it with data."""
    field_names = tech_data[0]._fields
    sheet = work_book.create_sheet(technology)

    for col_id, field in enumerate(field_names, start=1):
        sheet.cell(row=1, column=col_id, value=field)

    row = 2
    for cell in tech_data:
        for col, cell_parameter in enumerate(cell, start=1):
            sheet.cell(row=row, column=col, value=cell_parameter)
        row += 1


def make_excel_content(cell_data: AllTechPolygon) -> bytes:
    """Create an Excel file for different technologies and returns it as a BytesIO object."""
    work_book = Workbook()

    for tech, tech_data in cell_data.items():
        if tech == 'sites':
            continue
        _create_sheet(work_book, tech, tech_data)  # type: ignore

    work_book.remove(work_book['Sheet'])

    with NamedTemporaryFile() as tmp:
        work_book.save(tmp.name)
        tmp.seek(0)
        return tmp.read()
