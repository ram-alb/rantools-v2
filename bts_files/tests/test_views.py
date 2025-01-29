from http import HTTPStatus
from io import BytesIO
from unittest.mock import patch

import pytest
from django.urls import reverse_lazy
from openpyxl import load_workbook

from bts_files.services.atoll.gsm import GsmRowFactory

URL = reverse_lazy('bts-files')

lon_inside_udr = 78.36486111
lat_inside_udr = 45.0115

lon_outside_udr = 69.537921
lat_outside_udr = 46.94588

gsm_data = [
    GsmRowFactory(
        bsc='BSC1',
        site='GSite1',
        longitude=lon_inside_udr,
        latitude=lat_inside_udr,
        cell='GCell1',
        cid=1,
        bcch=10,
        bsic='01',
        lac='111',
        height=10,
        azimut=10,
        fband='GSM 900',
        antenna='Antenna1',
    ),
    GsmRowFactory(
        bsc='BSC1',
        site='GSite2',
        longitude=lon_outside_udr,
        latitude=lat_outside_udr,
        cell='GCell2',
        cid=1,
        bcch=10,
        bsic='01',
        lac='111',
        height=10,
        azimut=10,
        fband='GSM 900',
        antenna='Antenna2',
    ),
]

atoll_data = {'GSM': gsm_data}


def _is_contains_word(response, word: str) -> bool:
    content_type = response['Content-Type']
    if 'text' in content_type:
        file_content = response.content.decode()
        return word in file_content
    elif 'excel' in content_type:
        file_content = BytesIO(response.content)
        wb = load_workbook(filename=file_content)
        sheet = wb.active

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == word:
                    return True
    return False


def test_bts_files_get_rnpo_user(client, rnpo_user, get_templates):
    """Test a GET request to the BtsFiles view by an RNPO user."""
    client.login(
        username=rnpo_user['username'],
        password=rnpo_user['password'],
    )
    response = client.get(URL)

    assert response.status_code == HTTPStatus.OK
    assert 'bts_files/bts_files.html' in get_templates(response)


@pytest.mark.parametrize('file_type', ['kml', 'nbf', 'excel'])
def test_post_request(client, rnpo_user, file_type):
    """Test POST request to the BtsFiles view by an RNPO user returns the correct file content."""
    ext = 'xlsx' if file_type == 'excel' else file_type

    form_data = {
        'file-type': file_type,
        'technologies': ['GSM'],
        'regions': ['Jetysu-region'],
    }

    client.login(
        username=rnpo_user['username'],
        password=rnpo_user['password'],
    )

    with patch('bts_files.services.main.select_atoll_data', return_value=atoll_data):
        response = client.post(URL, form_data)

        assert response.status_code == HTTPStatus.OK
        assert response['Content-Disposition'].startswith(f'attachment; filename="bts_file.{ext}"')
        assert _is_contains_word(response, 'GSite1')
        assert not _is_contains_word(response, 'GSite2')
