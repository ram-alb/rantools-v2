from http import HTTPStatus
from io import BytesIO
from unittest.mock import patch

from django.urls import reverse_lazy
from openpyxl import load_workbook

URL = reverse_lazy("tr-data-index")

fake_tr_data = (
    ["header1", "header2"],
    [("data1", "data2"), ("data3", "data4")],
)


def _is_response_contains_word(response, word):
    file_content = BytesIO(response.content)
    wb = load_workbook(filename=file_content)

    sheet = wb["SITE_TR_DATA"]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == word:
                return True
    return False


@patch("tr_data.views.select_tr_data")
def test_post_request(mock_select_tr_data, client, rnpo_user):
    """Test post method of TrData view."""
    mock_select_tr_data.return_value = fake_tr_data
    file_name = "tr-data.xlsx"

    client.login(
        username=rnpo_user["username"],
        password=rnpo_user["password"],
    )
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response = client.post(URL, {"action": "ip"})

    assert response.status_code == HTTPStatus.OK
    assert response["Content-Type"] == content_type
    assert f'attachment; filename="{file_name}"' in response["Content-Disposition"]
    assert _is_response_contains_word(response, "data1")
    assert _is_response_contains_word(response, "data4")
