from typing import List

import openpyxl

from neighbors.services.excel import NeighborPair


def make_nonexistent_cells_report(nonexistent_cells: List[NeighborPair], date_time: str) -> str:
    """Generate an Excel report with information on nonexistent cells."""
    output_filename_path = f'neighbors/reports/nonexistent_cells_{date_time}.xlsx'

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Source Cell'  # type: ignore
    sheet['B1'] = 'Target Cell'  # type: ignore

    for row, nbr_pair in enumerate(nonexistent_cells, start=2):
        sheet.cell(row=row, column=1, value=nbr_pair.source_cell)  # type: ignore
        sheet.cell(row=row, column=2, value=nbr_pair.target_cell)  # type: ignore

    workbook.save(output_filename_path)
    workbook.close()

    return output_filename_path
