import zipfile
from io import BytesIO

from openpyxl import load_workbook


def _is_word_in_excel(zip_file, file_names, word):
    excel_files = [name for name in file_names if name.endswith('.xlsx')]
    excel_file_name = excel_files[0]

    with zip_file.open(excel_file_name) as excel_file:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == word:
                    return True
    return False


def _is_word_in_text(zip_file, file_names, word):
    text_files = [name for name in file_names if name.endswith('.xml') or name.endswith('.txt')]
    text_file_name = text_files[0]

    with zip_file.open(text_file_name) as text_file:
        text_content = text_file.read().decode('utf-8')
        return word in text_content


def is_word_in_zip(zip_archive, file_type, word):
    """Check for the presence of a word in files of a specified type within a zip archive."""
    with zipfile.ZipFile(BytesIO(zip_archive)) as zip_file:
        file_names = zip_file.namelist()

        if file_type == 'text':
            return _is_word_in_text(zip_file, file_names, word)

        elif file_type == 'excel':
            return _is_word_in_excel(zip_file, file_names, word)
