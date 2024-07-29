"""Create excel file with network live cell data."""
from tempfile import NamedTemporaryFile

from openpyxl import Workbook  # type: ignore

from network_live.services.select import NetworkLiveData


def create_excel(network_live_data: NetworkLiveData) -> bytes:
    """Create excel file with network live cell data."""
    work_book = Workbook()

    for technology, (cell_data, headers) in network_live_data.items():
        work_sheet = work_book.create_sheet(technology.upper())

        row = 1
        for column, header in enumerate(headers, start=1):
            work_sheet.cell(row=row, column=column, value=header)

        row = 2  # Start after headers
        for cell_values in cell_data:
            work_sheet.append(cell_values)
            row += 1

    work_book.remove(work_book['Sheet'])

    with NamedTemporaryFile() as tmp:
        work_book.save(tmp.name)
        tmp.seek(0)
        return tmp.read()
