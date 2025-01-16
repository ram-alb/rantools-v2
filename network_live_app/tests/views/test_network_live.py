from http import HTTPStatus
from io import BytesIO
from unittest.mock import patch

from django.urls import reverse_lazy
from openpyxl import load_workbook

URL = reverse_lazy('nl-index')

network_live_data = {
    'gsm': ([('data1', 'data2')], ['header1', 'header2']),
    'nr': ([('data3', 'data4')], ['header3', 'header4']),
}


def _is_contains_word(response, technologies, word):
    file_content = BytesIO(response.content)
    wb = load_workbook(filename=file_content)

    for tech in technologies:
        sheet = wb[tech.upper()]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == word:
                    return True
    return False


@patch('network_live_app.views.select_data')
def test_post_request(mock_select_data, client, rnpo_user):
    """Test post method of Network Live view."""
    mock_select_data.return_value = network_live_data

    form_data = {'technologies': ['gsm', 'nr']}
    file_name = 'nl_cells.xlsx'

    client.login(
        username=rnpo_user['username'],
        password=rnpo_user['password'],
    )
    response = client.post(URL, data=form_data)

    assert response.status_code == HTTPStatus.OK
    assert response['Content-Type'] == 'application/vnd.ms-excel'
    assert f'attachment; filename="{file_name}"' in response['Content-Disposition']
    assert _is_contains_word(response, form_data['technologies'], 'data1')
    assert _is_contains_word(response, form_data['technologies'], 'data4')
